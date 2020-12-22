"""
Before running these tests you must install the electrumsv-sdk and run:

electrumsv-sdk start node
electrumsv-sdk start electrumx
electrumsv-sdk start --new electrumsv

"""
import os
from pathlib import Path

import aiohttp
import asyncio
import json
import logging
from openpyxl import Workbook, load_workbook
import requests
import threading
import time

from electrumsv.restapi import Fault
from ..websocket_client import TxStateWSClient


MODULE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
logger = logging.getLogger("transaction_processing")


async def wait_for_mempool(txids):
    async with TxStateWSClient() as ws_client:
        await ws_client.block_until_mempool(txids)


async def wait_for_confirmation(txids):
    async with TxStateWSClient() as ws_client:
        await ws_client.block_until_confirmed(txids)


class TestRestAPI:

    TEST_WALLET_NAME = "worker1.sqlite"

    def _load_wallet(self):
        _result = requests.post(f'http://127.0.0.1:9999/v1/regtest/dapp/wallets/'
                                 f'{self.TEST_WALLET_NAME}/load_wallet')
        if _result.status_code != 200:
            raise requests.exceptions.HTTPError(_result.text)
        return _result

    def _topup_account(self):
        result = requests.post(f'http://127.0.0.1:9999/v1/regtest/dapp/wallets/'
                                f'worker1.sqlite/1/topup_account')
        if result.status_code != 200:
            raise requests.exceptions.HTTPError(result.text)
        return result

    def _generate_blocks(self, nblocks: int):
        payload = {"nblocks": nblocks}
        result = requests.post(f'http://127.0.0.1:9999/v1/regtest/dapp/wallets/'
                               f'{self.TEST_WALLET_NAME}/1/generate_blocks',
            data=json.dumps(payload))
        if result.status_code != 200:
            raise requests.exceptions.HTTPError(result.text)
        return result

    def _get_coin_state(self):
        result = requests.get(f'http://127.0.0.1:9999/v1/regtest/dapp/wallets/'
                               f'{self.TEST_WALLET_NAME}/1/utxos/coin_state')
        if result.status_code != 200:
            raise requests.exceptions.HTTPError(result.text)
        return result

    def _split_utxos(self, output_count: int=100, value: int=20000, desired_utxo_count: int=2000):
        SPLIT_FAILED_CODE = 40014
        payload = {
            "split_value": value,
            "split_count": output_count,
            "desired_utxo_count": desired_utxo_count,
            "password": "test"
        }
        url = f'http://127.0.0.1:9999/v1/regtest/dapp/wallets/' \
              f'{self.TEST_WALLET_NAME}/1/txs/split_utxos'

        result = requests.post(url, json=payload)
        if result.status_code != 200:
            if result.text:
                result_json = json.loads(result.text)
                if result_json.get('code') == SPLIT_FAILED_CODE:
                    return
            raise requests.exceptions.HTTPError(result.text)
        return result

    async def create_and_send_task(self, session, msg_queue, txid_queue):
        INSUFFICIENT_COINS_CODE = 40006
        while True:
            payload = await msg_queue.get()
            if not payload:  # poison pill
                break
            url = f'http://127.0.0.1:9999/v1/regtest/dapp/wallets/' \
                  f'{self.TEST_WALLET_NAME}/1/txs/create_and_broadcast'
            async with session.post(url, data=json.dumps(payload)) as resp:
                if resp.status != 200:
                    result = await resp.text()
                    logger.exception(f"tx creation and broadcast exception: {result}")
                    if result:
                        result_json = json.loads(result)
                    error_code = result_json.get('code')
                    if error_code == INSUFFICIENT_COINS_CODE:
                        await asyncio.sleep(0.2)
                        await msg_queue.put(payload)  # put back to queue and retry
                        continue
                    if error_code:
                        txid_queue.put_nowait(Fault(error_code, result_json.get('message')))
                        continue

                result_json = await resp.json()
                txid_queue.put_nowait(result_json['txid'])

    def autominer(self):
        while True:
            self._generate_blocks(1)
            time.sleep(10)

    def fill_queue(self, n_txs, msg_queue):
        # OP_RETURN OP_11 Hello World
        P2PKH_OUTPUT = {"script_pubkey": "006a0b68656c6c6f20776f726c64", "value": 0}
        payload = {
            "outputs": [P2PKH_OUTPUT],
            "password": "test"
        }

        # Fill queue with txs with this op_return output
        for i in range(n_txs):
            msg_queue.put_nowait(payload)

        logger.debug(f"filled msg_queue with {n_txs} payloads ready for tx creation and "
                     f"broadcast")

    async def wait_for_completion(self, txid_queue, n_txs, n_txn_creation_tasks, msg_queue):
        tx_count = 0
        txids = set()
        while tx_count < n_txs:
            txid = await txid_queue.get()
            if isinstance(txid, Fault):
                return False, txid
            logger.debug(f"got txid from queue {txid} - tx_count={tx_count}")
            txids.add(txid)
            tx_count += 1

        await wait_for_confirmation(txids)
        logger.debug(f"completed all block confirmation notifications (tx count = {n_txs})")

        for i in range(n_txn_creation_tasks):
            msg_queue.put_nowait(None)  # poison pill all tasks

        return True, None

    def test_concurrent_tx_creation_and_broadcast(self, event_loop):

        logging.basicConfig(level=logging.DEBUG)

        async def main():
            self._load_wallet()
            result = self._topup_account()
            self._generate_blocks(1)
            await wait_for_confirmation([result.json()['txid']])

            N_TXS = int(os.environ.get("STRESSTEST_N_TXS") or 2000)
            N_TX_CREATION_TASKS = 100
            DESIRED_UTXO_COUNT = int(os.environ.get("STRESSTEST_DESIRED_UTXO_COUNT") or 5000)
            SPLIT_TX_MAX_OUTPUTS = int(os.environ.get("STRESSTEST_SPLIT_TX_MAX_OUTPUTS") or 2000)
            msg_queue = asyncio.Queue()
            txid_queue = asyncio.Queue()

            # 1) split utxos
            confirmed_utxo_count = self._get_coin_state().json()['settled_coins']
            while confirmed_utxo_count < DESIRED_UTXO_COUNT:  # 10 x 1000 utxo splitting txs
                result_split = self._split_utxos(output_count=SPLIT_TX_MAX_OUTPUTS, value=10000,
                    desired_utxo_count=DESIRED_UTXO_COUNT)
                if not result_split:
                    break
                txid = result_split.json()['txid']
                self._generate_blocks(1)
                await wait_for_confirmation([txid])
                result = self._get_coin_state()
                confirmed_utxo_count = result.json()['settled_coins']
                logger.debug(f"confirmed_utxo_count={confirmed_utxo_count}")

            # Auto-mine regtest blocks
            t = threading.Thread(target=self.autominer, daemon=True)
            t.start()

            async with aiohttp.ClientSession() as session:
                tasks = [
                    asyncio.create_task(self.create_and_send_task(session, msg_queue, txid_queue))
                    for _ in
                         range(0, N_TX_CREATION_TASKS)
                ]

                t0 = time.time()

                self.fill_queue(N_TXS, msg_queue)
                passed, fault = await self.wait_for_completion(txid_queue, N_TXS,
                    N_TX_CREATION_TASKS, msg_queue)
                if not passed:
                    for task in tasks:
                        task.cancel()
                        try:
                            await task
                        except asyncio.CancelledError:
                            pass
                    assert False, fault
                t_delta = time.time() - t0

            results_filepath = MODULE_DIR.joinpath(".benchmarks/bench_result.xlsx")
            if not os.path.exists(results_filepath):
                wb = Workbook()
                ws = wb.active
                headers = ["desired_utxo_count", "outputs_per_splitting_tx", "n_txs", "total time",
                    "av. txn/sec"]
                ws.append(headers)
                wb.save(results_filepath)

            wb = load_workbook(results_filepath)
            ws = wb.active
            row = [DESIRED_UTXO_COUNT, SPLIT_TX_MAX_OUTPUTS, N_TXS, t_delta, N_TXS/t_delta]
            ws.append(row)
            wb.save(results_filepath)

        event_loop.run_until_complete(main())

